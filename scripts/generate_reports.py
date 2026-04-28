"""Generate all 7 reports from catches_scored.csv + anglers.csv + competitions.csv.

Outputs (under reports/):
  01_club_results_<comp>.xlsx
  02_details_of_fish_caught_<comp>.xlsx
  03_individual_position_in_club_<comp>.xlsx
  04_overall_club_results_<comp>.xlsx
  05_overall_individual_position_<comp>.xlsx
  06_overall_individual_position_per_league_<comp>.xlsx
  07_summary_of_fish_caught_<season>.xlsx

Scoring rule (confirmed 2026-04-28 by Shahiem):
  Edible fish     = 4 points per kg
  Non-edible fish = 1 point per kg
  Sub-team points = sum of angler points in that sub-team.
  Club points     = sum of all sub-team points in the club.

If non-edibles should score, update NON_EDIBLE_PTS_PER_KG below.
"""
from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
SEASONS = DATA / "seasons"
ACTIVE_FILE = DATA / "active_season.txt"


def _active_season() -> str:
    if ACTIVE_FILE.exists():
        s = ACTIVE_FILE.read_text(encoding="utf-8").strip()
        if s and (SEASONS / s).exists():
            return s
    seasons = sorted(p.name for p in SEASONS.iterdir() if p.is_dir()) if SEASONS.exists() else []
    return seasons[0] if seasons else "2025-26"


SEASON = _active_season()
SEASON_DIR = SEASONS / SEASON
REPORTS = ROOT / "reports" / SEASON
REPORTS.mkdir(parents=True, exist_ok=True)

FONT = "Arial"
HFONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
HFILL = PatternFill("solid", start_color="305496")
BFONT = Font(name=FONT, size=10)
TFONT = Font(name=FONT, bold=True, size=11)
TITLE = Font(name=FONT, bold=True, size=14)


# ---- Scoring rule ---------------------------------------------------------

import math

EDIBLE_PTS_PER_KG = 4.0
NON_EDIBLE_PTS_PER_KG = 1.0
NON_EDIBLE_MIN_KG = 1.0  # non-edible catches under this threshold score 0


def _floor2(x: float) -> float:
    return math.floor(float(x) * 100) / 100.0


def score_catch(weight_kg: float, edible: str) -> float:
    """Points awarded for a single catch — floored to 2 decimal places."""
    w = float(weight_kg or 0.0)
    is_edible = str(edible).upper() == "Y"
    if not is_edible and w < NON_EDIBLE_MIN_KG:
        return 0.00
    rate = EDIBLE_PTS_PER_KG if is_edible else NON_EDIBLE_PTS_PER_KG
    return _floor2(w * rate)


# ---- Data load ------------------------------------------------------------

def load() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    catches = pd.read_csv(SEASON_DIR / "catches_scored.csv")
    anglers = pd.read_csv(SEASON_DIR / "anglers.csv")
    comps = pd.read_csv(SEASON_DIR / "competitions.csv")
    catches["wp_no"] = catches["wp_no"].astype(str).str.strip()
    anglers["wp_no"] = anglers["wp_no"].astype(str).str.strip()
    catches["points"] = catches.apply(lambda r: score_catch(r["weight_kg"], r["edible"]), axis=1)
    return catches, anglers, comps


# ---- Excel helpers --------------------------------------------------------

def write_table(ws, df: pd.DataFrame, *, start_row: int = 1, title: str | None = None):
    r = start_row
    if title:
        ws.cell(row=r, column=1, value=title).font = TITLE
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max(1, len(df.columns)))
        r += 2
    for c, col in enumerate(df.columns, 1):
        cell = ws.cell(row=r, column=c, value=str(col))
        cell.font = HFONT; cell.fill = HFILL
        cell.alignment = Alignment(horizontal="center")
    for i, row in enumerate(df.itertuples(index=False), r + 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=c, value=val if pd.notna(val) else None)
            cell.font = BFONT
    for c in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(c)
        max_len = max([len(str(df.columns[c - 1]))] + [len(str(v)) for v in df.iloc[:, c - 1].fillna("")])
        ws.column_dimensions[col_letter].width = min(40, max(10, max_len + 2))


def save(wb: Workbook, name: str) -> Path:
    path = REPORTS / name
    wb.save(path)
    return path


# ---- Helpers --------------------------------------------------------------

def comp_catches(catches: pd.DataFrame, comp_id: str) -> pd.DataFrame:
    return catches[catches["comp_id"].astype(str).str.strip() == comp_id].copy()


def angler_full_name(a: pd.Series) -> str:
    return f"{a['first_name']} {a['surname']}".strip()


# ---- Reports --------------------------------------------------------------

def report_01_club_results(catches, anglers, comp_id: str):
    cc = comp_catches(catches, comp_id).merge(
        anglers[["wp_no", "club", "sub_team"]], on="wp_no", how="left"
    )
    cc["sub_team"] = cc["sub_team"].fillna("").replace("", "A")
    pivot = cc.pivot_table(index="club", columns="sub_team",
                           values="points", aggfunc="sum", fill_value=0)
    pivot.columns = [f"PNTS_{c}" for c in pivot.columns]
    pivot["TOTAL"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("TOTAL", ascending=False).reset_index()
    pivot.insert(0, "RANK", range(1, len(pivot) + 1))
    wb = Workbook(); ws = wb.active; ws.title = "Club Results"
    write_table(ws, pivot, title=f"Club Results — {comp_id}")
    return save(wb, f"01_club_results_{comp_id.replace(' ', '_')}.xlsx")


def report_02_details_of_fish_caught(catches, anglers, comp_id: str):
    cc = comp_catches(catches, comp_id).merge(anglers, on="wp_no", how="left")
    cc["angler"] = cc.apply(lambda r: f"{r['first_name']} {r['surname']}" if pd.notna(r['first_name']) else "(unknown)", axis=1)
    out = cc[["club", "wp_no", "angler", "species_raw", "weight_kg", "length_cm", "edible"]] \
        .rename(columns={"species_raw": "species", "weight_kg": "weight (kg)", "length_cm": "length (cm)"}) \
        .sort_values(["club", "wp_no", "species"]).reset_index(drop=True)
    wb = Workbook(); ws = wb.active; ws.title = "Details"
    write_table(ws, out, title=f"Details of Fish Caught — {comp_id}")
    return save(wb, f"02_details_of_fish_caught_{comp_id.replace(' ', '_')}.xlsx")


def report_03_individual_position_in_club(catches, anglers, comp_id: str):
    cc = comp_catches(catches, comp_id)
    by_angler = cc.groupby("wp_no", as_index=False).agg(
        catches=("species_raw", "count"),
        total_weight=("weight_kg", "sum"),
        points=("points", "sum"),
    )
    out = anglers.merge(by_angler, on="wp_no", how="left").fillna({"catches": 0, "total_weight": 0, "points": 0})
    out["angler"] = out["first_name"] + " " + out["surname"]
    out = out[["club", "sub_team", "wp_no", "angler", "league_code",
               "catches", "total_weight", "points"]]
    wb = Workbook()
    for club, g in out.groupby("club"):
        ws = wb.create_sheet(str(club)[:31])
        g = g.sort_values(["sub_team", "points"], ascending=[True, False]).reset_index(drop=True)
        g.insert(0, "rank", g.groupby("sub_team")["points"].rank(method="dense", ascending=False).fillna(0).astype(int))
        write_table(ws, g, title=f"Individual Position in Club — {club} — {comp_id}")
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    return save(wb, f"03_individual_position_in_club_{comp_id.replace(' ', '_')}.xlsx")


def report_04_overall_club_results(catches, anglers, through_comp: str):
    comp_order = sorted(catches["comp_id"].astype(str).str.strip().unique())
    if through_comp in comp_order:
        comp_order = comp_order[: comp_order.index(through_comp) + 1]
    cc = catches[catches["comp_id"].astype(str).str.strip().isin(comp_order)].merge(
        anglers[["wp_no", "club"]], on="wp_no", how="left"
    )
    pivot = cc.pivot_table(index="club", columns="comp_id",
                           values="points", aggfunc="sum", fill_value=0)
    pivot["TOTAL"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("TOTAL", ascending=False).reset_index()
    pivot.insert(0, "RANK", range(1, len(pivot) + 1))
    wb = Workbook(); ws = wb.active; ws.title = "Overall Club"
    write_table(ws, pivot, title=f"Overall Club Results — through {through_comp}")
    return save(wb, f"04_overall_club_results_{through_comp.replace(' ', '_')}.xlsx")


def report_05_overall_individual_position(catches, anglers, through_comp: str):
    comp_order = sorted(catches["comp_id"].astype(str).str.strip().unique())
    if through_comp in comp_order:
        comp_order = comp_order[: comp_order.index(through_comp) + 1]
    cc = catches[catches["comp_id"].astype(str).str.strip().isin(comp_order)]
    pivot = cc.pivot_table(index="wp_no", columns="comp_id",
                           values="points", aggfunc="sum", fill_value=0).reset_index()
    out = anglers.merge(pivot, on="wp_no", how="left").fillna(0)
    out["angler"] = out["first_name"] + " " + out["surname"]
    comp_cols = [c for c in comp_order if c in out.columns]
    out["TOTAL"] = out[comp_cols].sum(axis=1) if comp_cols else 0
    out = out[["club", "wp_no", "angler", "league_code"] + comp_cols + ["TOTAL"]]
    out = out.sort_values("TOTAL", ascending=False).reset_index(drop=True)
    out.insert(0, "RANK", range(1, len(out) + 1))
    wb = Workbook(); ws = wb.active; ws.title = "Overall Individual"
    write_table(ws, out, title=f"Overall Individual Position — through {through_comp}")
    return save(wb, f"05_overall_individual_position_{through_comp.replace(' ', '_')}.xlsx")


def report_06_overall_individual_per_league(catches, anglers, through_comp: str):
    p = report_05_overall_individual_position  # build same data, then split by league
    comp_order = sorted(catches["comp_id"].astype(str).str.strip().unique())
    if through_comp in comp_order:
        comp_order = comp_order[: comp_order.index(through_comp) + 1]
    cc = catches[catches["comp_id"].astype(str).str.strip().isin(comp_order)]
    pivot = cc.pivot_table(index="wp_no", columns="comp_id",
                           values="points", aggfunc="sum", fill_value=0).reset_index()
    out = anglers.merge(pivot, on="wp_no", how="left").fillna(0)
    out["angler"] = out["first_name"] + " " + out["surname"]
    comp_cols = [c for c in comp_order if c in out.columns]
    out["TOTAL"] = out[comp_cols].sum(axis=1) if comp_cols else 0
    out = out[["league_code", "club", "wp_no", "angler"] + comp_cols + ["TOTAL"]]
    wb = Workbook()
    for lg, g in out.groupby("league_code"):
        ws = wb.create_sheet(f"League {lg}")
        g = g.sort_values("TOTAL", ascending=False).reset_index(drop=True)
        g.insert(0, "RANK", range(1, len(g) + 1))
        write_table(ws, g, title=f"League {lg} — through {through_comp}")
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    return save(wb, f"06_overall_individual_position_per_league_{through_comp.replace(' ', '_')}.xlsx")


def report_07_summary_of_fish_caught(catches, season_label: str):
    comp_order = sorted(catches["comp_id"].astype(str).str.strip().unique())
    pivot = catches.pivot_table(
        index="species_raw", columns="comp_id",
        values="weight_kg", aggfunc="count", fill_value=0,
    )
    pivot = pivot.reindex(columns=comp_order, fill_value=0)
    pivot["TOTAL"] = pivot.sum(axis=1)
    pivot = pivot.reset_index().sort_values("species_raw")
    weight_per_comp = catches.groupby("comp_id")["weight_kg"].sum().reindex(comp_order, fill_value=0)
    fish_per_comp = catches.groupby("comp_id")["species_raw"].count().reindex(comp_order, fill_value=0)
    edible_split = catches.groupby(["comp_id", "edible"])["species_raw"].count().unstack(fill_value=0).reindex(comp_order, fill_value=0)

    wb = Workbook(); ws = wb.active; ws.title = "Species x Comp"
    write_table(ws, pivot, title=f"Summary of Fish Caught — {season_label}")
    ws2 = wb.create_sheet("Totals per Comp")
    totals = pd.DataFrame({"comp_id": comp_order,
                           "fish_count": fish_per_comp.values,
                           "total_weight_kg": weight_per_comp.values})
    totals = totals.merge(edible_split.reset_index(), on="comp_id", how="left").fillna(0)
    write_table(ws2, totals, title="Totals per Competition")
    return save(wb, f"07_summary_of_fish_caught_{season_label}.xlsx")


# ---- CLI ------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--comp", required=True, help="Competition id, e.g. 'IC 8'")
    ap.add_argument("--season", default=SEASON)
    args = ap.parse_args()

    catches, anglers, comps = load()
    paths = [
        report_01_club_results(catches, anglers, args.comp),
        report_02_details_of_fish_caught(catches, anglers, args.comp),
        report_03_individual_position_in_club(catches, anglers, args.comp),
        report_04_overall_club_results(catches, anglers, args.comp),
        report_05_overall_individual_position(catches, anglers, args.comp),
        report_06_overall_individual_per_league(catches, anglers, args.comp),
        report_07_summary_of_fish_caught(catches, args.season),
    ]
    for p in paths:
        print(f"  wrote {p}")


if __name__ == "__main__":
    main()
