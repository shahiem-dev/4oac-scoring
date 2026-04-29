"""Build the WCSAA IC League Tracker — a single print-ready XLSX with:

  Dashboard          — club ranking + top 20 individuals (snapshot)
  Club Standings     — clubs x IC matrix with totals + rank
  Individual Standings — every angler ranked by season total
  <one sheet per club> — anglers in that club, points per IC + total + rank
  Notes              — scoring rule + how-to-refresh

Refresh by re-running:
    python scripts/score_catches.py
    python scripts/build_tracker.py

Pages are configured for landscape A4, fit-to-width, with the header
row repeated on every printed page.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.worksheet import Worksheet

from generate_reports import score_catch  # reuse the league scoring rule

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
SEASONS = DATA / "seasons"
ACTIVE_FILE = DATA / "active_season.txt"


def _active_season() -> str:
    if ACTIVE_FILE.exists():
        s = ACTIVE_FILE.read_text(encoding="utf-8").strip()
        if s and (SEASONS / s).exists():
            return s
    seasons = sorted(p.name for p in SEASONS.iterdir() if p.is_dir()) if SEASONS.exists() else []
    return seasons[0] if seasons else "2025-26"


SEASON = _active_season()
SEASON_DIR = SEASONS / SEASON

_DESKTOP_CANDIDATES = [
    Path.home() / "OneDrive - Africa Cricket Development (Pty) Ltd" / "Desktop",
    Path.home() / "OneDrive" / "Desktop",
    Path.home() / "Desktop",
]
DESKTOP = next((p for p in _DESKTOP_CANDIDATES if p.exists()), _DESKTOP_CANDIDATES[-1])
OUT = DESKTOP / "League" / f"WCSAA_League_Tracker_{SEASON}.xlsx"

FONT = "Arial"
TITLE = Font(name=FONT, bold=True, size=16, color="FFFFFF")
TITLE_FILL = PatternFill("solid", start_color="1F4E78")
SUB = Font(name=FONT, italic=True, size=10, color="595959")
HFONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
HFILL = PatternFill("solid", start_color="305496")
BFONT = Font(name=FONT, size=10)
TOTAL_FONT = Font(name=FONT, bold=True, size=10)
TOTAL_FILL = PatternFill("solid", start_color="DDEBF7")
ALT_FILL = PatternFill("solid", start_color="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def _load_team_assignments(season_dir):
    p = season_dir / "team_assignments.csv"
    if not p.exists():
        return pd.DataFrame(columns=["comp_id", "wp_no", "sub_team"])
    df = pd.read_csv(p, dtype=str).fillna("")
    df["comp_id"] = df["comp_id"].str.strip()
    df["wp_no"] = df["wp_no"].str.strip()
    df["sub_team"] = df["sub_team"].str.strip().str.upper()
    return df


def _attach_sub_team(catches, anglers, season_dir):
    ta = _load_team_assignments(season_dir)
    out = catches.copy()
    if not ta.empty:
        out = out.merge(ta.rename(columns={"sub_team": "_a"}),
                        on=["comp_id", "wp_no"], how="left")
    else:
        out["_a"] = ""
    out = out.merge(anglers[["wp_no", "sub_team"]].rename(columns={"sub_team": "_d"}),
                    on="wp_no", how="left")
    a = out["_a"].fillna("").astype(str).str.upper().str.strip()
    d = out["_d"].fillna("").astype(str).str.upper().str.strip()
    out["sub_team"] = a.where(a != "", d)
    return out.drop(columns=["_a", "_d"], errors="ignore")


def load() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, list[str]]:
    catches = pd.read_csv(SEASON_DIR / "catches_scored.csv")
    anglers = pd.read_csv(SEASON_DIR / "anglers.csv")
    comps = pd.read_csv(SEASON_DIR / "competitions.csv")
    catches["wp_no"] = catches["wp_no"].astype(str).str.strip()
    anglers["wp_no"] = anglers["wp_no"].astype(str).str.strip()
    catches["comp_id"] = catches["comp_id"].astype(str).str.strip()
    catches["points"] = catches.apply(lambda r: score_catch(r["weight_kg"], r["edible"]), axis=1)
    catches = _attach_sub_team(catches, anglers, SEASON_DIR)
    comp_order = sorted(catches["comp_id"].unique().tolist())
    return catches, anglers, comps, comp_order


def setup_print(ws: Worksheet, *, header_rows: int = 2, repeat_header_row: int | None = None) -> None:
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6, header=0.3, footer=0.3)
    ws.print_options.horizontalCentered = True
    if repeat_header_row:
        ws.print_title_rows = f"{repeat_header_row}:{repeat_header_row}"
    ws.oddFooter.center.text = "&P of &N"
    ws.oddFooter.right.text = "WCSAA League Tracker"


def write_title(ws: Worksheet, title: str, subtitle: str, ncols: int) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE; c.fill = TITLE_FILL; c.alignment = CENTER
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = SUB; c.alignment = CENTER
    return 4  # next free row


def write_table(ws: Worksheet, df: pd.DataFrame, *, start_row: int,
                total_row: bool = False, num_cols: list[str] | None = None) -> int:
    num_cols = num_cols or []
    # Header
    for c, col in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=c, value=str(col))
        cell.font = HFONT; cell.fill = HFILL; cell.alignment = CENTER; cell.border = BORDER
    ws.row_dimensions[start_row].height = 22
    # Body
    for i, row in enumerate(df.itertuples(index=False), 1):
        r = start_row + i
        fill = ALT_FILL if i % 2 == 0 else None
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val if pd.notna(val) else None)
            cell.font = BFONT; cell.border = BORDER
            cell.alignment = CENTER if df.columns[c - 1] in num_cols or c == 1 else LEFT
            if fill: cell.fill = fill
            if df.columns[c - 1] in num_cols and isinstance(val, (int, float)):
                cell.number_format = "#,##0.00" if isinstance(val, float) else "#,##0"
    last_row = start_row + len(df)
    if total_row and len(df):
        r = last_row + 1
        ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
        for c, col in enumerate(df.columns, 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = TOTAL_FILL; cell.border = BORDER; cell.font = TOTAL_FONT
            cell.alignment = CENTER if c > 1 else LEFT
            if col in num_cols:
                col_letter = get_column_letter(c)
                cell.value = f"=SUM({col_letter}{start_row + 1}:{col_letter}{last_row})"
                cell.number_format = "#,##0.00"
        last_row = r
    # Column widths
    for c, col in enumerate(df.columns, 1):
        col_letter = get_column_letter(c)
        max_len = max([len(str(col))] + [len(str(v)) for v in df.iloc[:, c - 1].fillna("")])
        ws.column_dimensions[col_letter].width = min(34, max(9, max_len + 2))
    return last_row


# ---- Builders -------------------------------------------------------------

def build_dashboard(wb: Workbook, catches, anglers, comp_order, season: str) -> None:
    ws = wb.create_sheet("Dashboard")
    cc = catches.merge(anglers[["wp_no", "club", "first_name", "surname", "league_code"]],
                       on="wp_no", how="left")
    cc["club"] = cc["club"].fillna("UNKNOWN")

    # Club totals
    club_tot = cc.groupby("club", as_index=False)["points"].sum() \
                 .sort_values("points", ascending=False).reset_index(drop=True)
    club_tot.insert(0, "Rank", range(1, len(club_tot) + 1))
    club_tot = club_tot.rename(columns={"club": "Club", "points": "Total Points"})

    # Top 20 individuals
    cc["angler"] = (cc["first_name"].fillna("") + " " + cc["surname"].fillna("")).str.strip().replace("", "(unknown)")
    ind = cc.groupby(["wp_no", "angler", "club", "league_code"], as_index=False, dropna=False)["points"].sum()
    ind = ind.sort_values("points", ascending=False).head(20).reset_index(drop=True)
    ind.insert(0, "Rank", range(1, len(ind) + 1))
    ind = ind.rename(columns={"wp_no": "WP No", "angler": "Angler", "club": "Club",
                              "league_code": "Lg", "points": "Total Points"})

    next_row = write_title(ws, "WCSAA League Dashboard", f"Season {season} — through {comp_order[-1] if comp_order else 'no comps'}",
                           ncols=6)
    ws.cell(row=next_row, column=1, value="Club Standings").font = Font(name=FONT, bold=True, size=12)
    next_row += 1
    last = write_table(ws, club_tot, start_row=next_row, total_row=True, num_cols=["Total Points"])
    next_row = last + 3
    ws.cell(row=next_row, column=1, value="Top 20 Individuals").font = Font(name=FONT, bold=True, size=12)
    next_row += 1
    write_table(ws, ind, start_row=next_row, num_cols=["Total Points"])

    setup_print(ws)
    ws.sheet_view.showGridLines = False
    wb.active = wb.sheetnames.index("Dashboard")


def build_club_standings(wb: Workbook, catches, anglers, comp_order) -> None:
    """Two views: per-comp totals (clubs × IC) and current sub-team breakdown (PNTS A..I + A+B)."""
    sub_teams = list("ABCDEFGHI")

    ws = wb.create_sheet("Club Standings")
    cc = catches.merge(anglers[["wp_no", "club"]], on="wp_no", how="left")
    cc["club"] = cc["club"].fillna("UNKNOWN")
    cc["sub_team"] = cc["sub_team"].fillna("").str.upper().str.strip()

    # ----- Per-comp matrix (clubs x IC) -----
    pivot = cc.pivot_table(index="club", columns="comp_id", values="points",
                           aggfunc="sum", fill_value=0)
    pivot = pivot.reindex(columns=comp_order, fill_value=0)
    pivot["Total"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("Total", ascending=False).reset_index()
    pivot.insert(0, "Pos.", range(1, len(pivot) + 1))
    pivot = pivot.rename(columns={"club": "Club"})

    next_row = write_title(ws, "Club Standings — All Competitions",
                           f"Through {comp_order[-1] if comp_order else 'no comps'}",
                           ncols=max(len(pivot.columns), 12))
    last = write_table(ws, pivot, start_row=next_row, total_row=True,
                       num_cols=list(comp_order) + ["Total"])

    # ----- Sub-team breakdown (PNTS A..I + A+B) -----
    next_row = last + 3
    ws.cell(row=next_row, column=1, value="Sub-team Breakdown (Season)").font = \
        Font(name=FONT, bold=True, size=12)
    next_row += 1
    sub_pivot = cc.pivot_table(index="club", columns="sub_team",
                               values="points", aggfunc="sum")
    sub_pivot = sub_pivot.reindex(columns=sub_teams)
    out = pd.DataFrame(index=sub_pivot.index)
    out["PNTS A"] = sub_pivot["A"]
    out["PNTS B"] = sub_pivot["B"]
    ab = sub_pivot["A"].fillna(0) + sub_pivot["B"].fillna(0)
    ab[sub_pivot["A"].isna() & sub_pivot["B"].isna()] = pd.NA
    out["PNTS A+B"] = ab
    for t in sub_teams[2:]:
        out[f"PNTS {t}"] = sub_pivot[t]
    out = out.sort_values("PNTS A+B", ascending=False, na_position="last").reset_index()
    out.insert(0, "Pos.", range(1, len(out) + 1))
    out = out.rename(columns={"club": "CLUB"})
    write_table(ws, out, start_row=next_row, total_row=False,
                num_cols=[c for c in out.columns if c.startswith("PNTS")])

    setup_print(ws, repeat_header_row=4)
    ws.sheet_view.showGridLines = False


def build_individual_standings(wb: Workbook, catches, anglers, comp_order) -> None:
    ws = wb.create_sheet("Individual Standings")
    cc = catches.merge(anglers[["wp_no", "club", "first_name", "surname", "league_code"]],
                       on="wp_no", how="left")
    cc["angler"] = (cc["first_name"].fillna("") + " " + cc["surname"].fillna("")).str.strip().replace("", "(unknown)")
    pivot = cc.pivot_table(index=["wp_no", "angler", "club", "league_code"],
                           columns="comp_id", values="points", aggfunc="sum",
                           fill_value=0).reset_index()
    pivot = pivot.reindex(columns=["wp_no", "angler", "club", "league_code"] + comp_order, fill_value=0)
    pivot["Total"] = pivot[comp_order].sum(axis=1) if comp_order else 0
    pivot = pivot.sort_values("Total", ascending=False).reset_index(drop=True)
    pivot.insert(0, "Rank", range(1, len(pivot) + 1))
    pivot = pivot.rename(columns={"wp_no": "WP No", "angler": "Angler",
                                  "club": "Club", "league_code": "Lg"})

    next_row = write_title(ws, "Individual Standings — Season",
                           f"Through {comp_order[-1] if comp_order else 'no comps'}",
                           ncols=len(pivot.columns))
    write_table(ws, pivot, start_row=next_row, num_cols=list(comp_order) + ["Total"])
    setup_print(ws, repeat_header_row=next_row)
    ws.sheet_view.showGridLines = False


def build_club_sheet(wb: Workbook, club: str, catches, anglers, comp_order) -> None:
    sheet_name = club[:31] if club else "UNKNOWN"
    ws = wb.create_sheet(sheet_name)
    club_anglers = anglers[anglers["club"].fillna("UNKNOWN") == club].copy()
    club_anglers["sub_team"] = club_anglers["sub_team"].fillna("").astype(str).str.upper().str.strip()
    cc = catches.merge(anglers[["wp_no", "club"]], on="wp_no", how="left")
    cc["club"] = cc["club"].fillna("UNKNOWN")
    cc["sub_team"] = cc["sub_team"].fillna("").astype(str).str.upper().str.strip()
    cc = cc[cc["club"] == club]

    # Per-angler points pivot
    pts = cc.pivot_table(index="wp_no", columns="comp_id", values="points",
                         aggfunc="sum", fill_value=0).reset_index()
    out = club_anglers.merge(pts, on="wp_no", how="left").fillna(0)
    out["angler"] = out["first_name"] + " " + out["surname"]
    for c in comp_order:
        if c not in out.columns:
            out[c] = 0
    out["Total"] = out[comp_order].sum(axis=1) if comp_order else 0
    out = out[["sub_team", "wp_no", "angler", "league_code"] + comp_order + ["Total"]]
    # Rank within sub-team, then overall sort
    out["Pos."] = (out.groupby("sub_team")["Total"]
                   .rank(method="dense", ascending=False).fillna(0).astype(int))
    out = out.sort_values(["sub_team", "Total"], ascending=[True, False]).reset_index(drop=True)
    out = out[["Pos.", "sub_team", "wp_no", "angler", "league_code"] + comp_order + ["Total"]]
    out = out.rename(columns={"sub_team": "Team", "wp_no": "WP No",
                              "angler": "Angler", "league_code": "Div"})

    next_row = write_title(ws, f"{club} — Members", f"Through {comp_order[-1] if comp_order else 'no comps'}",
                           ncols=len(out.columns))

    # Sub-team summary (A..I + A+B)
    sub_teams = list("ABCDEFGHI")
    sub_pivot = cc.groupby("sub_team")["points"].sum().reindex(sub_teams)
    summary = pd.DataFrame({
        "Team": sub_teams + ["A+B"],
        "Points": list(sub_pivot.values) + [
            (sub_pivot.get("A", 0) or 0) + (sub_pivot.get("B", 0) or 0)
            if not (pd.isna(sub_pivot.get("A")) and pd.isna(sub_pivot.get("B"))) else pd.NA
        ],
    })
    ws.cell(row=next_row, column=1, value="Sub-team Totals").font = Font(name=FONT, bold=True, size=12)
    next_row += 1
    last = write_table(ws, summary, start_row=next_row, num_cols=["Points"])
    next_row = last + 3
    ws.cell(row=next_row, column=1, value="Members (ranked within sub-team)").font = \
        Font(name=FONT, bold=True, size=12)
    next_row += 1
    last = write_table(ws, out, start_row=next_row, total_row=True,
                       num_cols=list(comp_order) + ["Total"])

    # Catch detail block
    detail = cc.merge(anglers[["wp_no", "first_name", "surname"]], on="wp_no", how="left")
    detail["Angler"] = detail["first_name"].fillna("") + " " + detail["surname"].fillna("")
    detail = detail[["comp_id", "wp_no", "Angler", "species_raw", "length_cm",
                     "weight_kg", "edible", "points", "status"]] \
        .rename(columns={"comp_id": "Comp", "wp_no": "WP No", "species_raw": "Species",
                         "length_cm": "Length (cm)", "weight_kg": "Weight (kg)",
                         "edible": "Edible", "points": "Points", "status": "Status"}) \
        .sort_values(["Comp", "WP No"]).reset_index(drop=True)
    if len(detail):
        next_row = last + 3
        ws.cell(row=next_row, column=1, value="Catch Detail").font = Font(name=FONT, bold=True, size=12)
        next_row += 1
        write_table(ws, detail, start_row=next_row,
                    num_cols=["Length (cm)", "Weight (kg)", "Points"])

    setup_print(ws, repeat_header_row=4)
    ws.sheet_view.showGridLines = False


def build_notes(wb: Workbook, season: str) -> None:
    ws = wb.create_sheet("Notes")
    notes = [
        ("WCSAA IC League Tracker", TITLE_FILL),
        (f"Season: {season}", None),
        ("", None),
        ("Scoring rule:", None),
        ("    Edible fish      = 4 points per kg", None),
        ("    Non-edible fish  = 1 point per kg (under 1.00 kg = 0 pts)", None),
        ("    All scores are floored to 2 decimal places.", None),
        ("    Site Fish (...)  = 1.00 kg flat (then × points-per-kg above)", None),
        ("    < X kg suffix    = 0 points (sub-minimum / participation)", None),
        ("    Catshark (Brown), Catshark (Puffadder) = 0 points", None),
        ("", None),
        ("Weight formula: W_kg = exp(log_a + b × ln(length_cm))   (SASAA)", None),
        ("", None),
        ("How to refresh:", None),
        ("    1. Open data\\catch_entry_template.xlsx → enter catches on the Catches sheet.", None),
        ("    2. python scripts\\score_catches.py     (computes weight + points)", None),
        ("    3. python scripts\\build_tracker.py     (regenerates this workbook)", None),
        ("    4. python scripts\\generate_reports.py  (per-comp printable reports)", None),
        ("", None),
        ("Print: every sheet is configured for landscape A4, fit-to-width, header row repeats.", None),
        ("       File → Print → Print Entire Workbook → Save as PDF.", None),
    ]
    for i, (text, fill) in enumerate(notes, 1):
        c = ws.cell(row=i, column=1, value=text)
        if i == 1:
            c.font = TITLE
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
            c.fill = TITLE_FILL; c.alignment = CENTER
            ws.row_dimensions[1].height = 26
        else:
            c.font = Font(name=FONT, size=11, bold=text.endswith(":"))
    ws.column_dimensions["A"].width = 100
    setup_print(ws)
    ws.sheet_view.showGridLines = False


# ---- Main -----------------------------------------------------------------

def main() -> None:
    catches, anglers, comps, comp_order = load()
    season = SEASON

    wb = Workbook()
    wb.remove(wb.active)

    build_dashboard(wb, catches, anglers, comp_order, season)
    build_club_standings(wb, catches, anglers, comp_order)
    build_individual_standings(wb, catches, anglers, comp_order)

    clubs = sorted(anglers["club"].fillna("UNKNOWN").unique().tolist())
    for club in clubs:
        build_club_sheet(wb, club, catches, anglers, comp_order)

    build_notes(wb, season)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"  Sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
