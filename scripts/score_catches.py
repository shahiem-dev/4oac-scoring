"""Read raw catches from catch_entry_template.xlsx, score them, write back
with computed columns filled in. Also writes catches_scored.csv for downstream
report generation.
"""
import math
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scoring import Scorer


def _floor2(x: float) -> float:
    return math.floor(float(x) * 100) / 100.0

DATA = Path(__file__).resolve().parent.parent / "data"
SEASONS = DATA / "seasons"
ACTIVE_FILE = DATA / "active_season.txt"


def _active_season() -> str:
    if ACTIVE_FILE.exists():
        s = ACTIVE_FILE.read_text(encoding="utf-8").strip()
        if s and (SEASONS / s).exists():
            return s
    seasons = sorted(p.name for p in SEASONS.iterdir() if p.is_dir()) if SEASONS.exists() else []
    return seasons[0] if seasons else "2025-26"


SEASON_DIR = SEASONS / _active_season()
TEMPLATE = DATA / "catch_entry_template.xlsx"
OUT_CSV = SEASON_DIR / "catches_scored.csv"


def main() -> None:
    s = Scorer()
    wb = load_workbook(TEMPLATE)
    ws = wb["Catches"]
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        comp_id, wp_no, species_raw, length_cm = (c.value for c in row[:4])
        if not species_raw or not wp_no or not comp_id:
            continue
        try:
            L = float(length_cm) if length_cm not in (None, "") else None
        except (TypeError, ValueError):
            L = None
        res = s.score(str(species_raw), L)
        row[4].value = res.canonical_name
        row[5].value = _floor2(res.weight_kg)
        row[6].value = res.edible
        row[7].value = res.note
        rows.append({
            "comp_id": comp_id, "wp_no": str(wp_no).strip(),
            "species_raw": species_raw,
            "canonical_species": res.canonical_name,
            "length_cm": L, "weight_kg": res.weight_kg,
            "edible": res.edible, "status": res.note,
        })
    wb.save(TEMPLATE)
    df = pd.DataFrame(rows)
    df.to_csv(OUT_CSV, index=False)
    print(f"Scored {len(df)} catches -> {OUT_CSV}")
    if len(df):
        unknown = df[df["status"] == "error:unknown_species"]
        if len(unknown):
            print(f"\n  WARNING: {len(unknown)} unknown species — will score 0:")
            for sp in unknown["species_raw"].unique():
                print(f"    - {sp}")


if __name__ == "__main__":
    main()
