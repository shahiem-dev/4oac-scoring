"""Build the catch-entry XLSX template (data\\catch_entry_template.xlsx).

Sheets:
  Anglers     — master roster (one row per angler)
  Competitions — IC 1..N with date + venue
  Catches     — entry sheet: comp_id, wp_no, species_raw, length_cm + auto-computed cols
  Reference   — read-only species master + alias map for data validation
"""
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

DATA = Path(__file__).resolve().parent.parent / "data"
OUT = DATA / "catch_entry_template.xlsx"

FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="305496")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name=FONT, size=10)
BLUE_INPUT = Font(name=FONT, size=10, color="0000FF")
BLACK_FORMULA = Font(name=FONT, size=10, color="000000")


def write_sheet(ws, df: pd.DataFrame, *, table_name: str, input_cols: set[str] | None = None):
    input_cols = input_cols or set()
    for c, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r, row in enumerate(df.itertuples(index=False), 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val if val == val else None)  # NaN -> None
            cell.font = BLUE_INPUT if df.columns[c - 1] in input_cols else BODY_FONT
    end_col = get_column_letter(len(df.columns))
    end_row = max(2, len(df) + 1)
    tbl = Table(displayName=table_name, ref=f"A1:{end_col}{end_row}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)
    for c in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(c)].width = max(14, min(40, df.columns[c - 1].__len__() + 4))


def main() -> None:
    anglers = pd.read_csv(DATA / "anglers.csv")
    comps = pd.read_csv(DATA / "competitions.csv")
    species = pd.read_csv(DATA / "species_master.csv")

    wb = Workbook()
    wb.remove(wb.active)

    # 1. Anglers sheet (input)
    ws = wb.create_sheet("Anglers")
    write_sheet(
        ws, anglers,
        table_name="tblAnglers",
        input_cols={"wp_no", "sasaa_no", "first_name", "surname", "club",
                    "sub_team", "league_division", "league_code"},
    )

    # 2. Competitions sheet (input)
    ws = wb.create_sheet("Competitions")
    write_sheet(ws, comps, table_name="tblComps",
                input_cols={"comp_id", "date", "venue"})

    # 3. Catches sheet — input cols + computed columns via lookups.
    # Calculations are intentionally done in Python (scripts/score_catches.py)
    # to honour the alias map + Site Fish rule + sub-minimum rule which are
    # cumbersome in pure Excel. Catches sheet only collects raw entries.
    catches_cols = ["comp_id", "wp_no", "species_raw", "length_cm",
                    "_canonical_species", "_weight_kg", "_edible", "_status"]
    df_catches = pd.DataFrame(columns=catches_cols)
    ws = wb.create_sheet("Catches")
    write_sheet(ws, df_catches, table_name="tblCatches",
                input_cols={"comp_id", "wp_no", "species_raw", "length_cm"})
    # Add 200 blank rows for entry
    for r in range(2, 202):
        for c in range(1, 5):
            ws.cell(row=r, column=c).font = BLUE_INPUT
        for c in range(5, 9):
            ws.cell(row=r, column=c).font = BLACK_FORMULA
    # Resize table to include blank rows
    tbl = ws.tables["tblCatches"]
    tbl.ref = f"A1:{get_column_letter(len(catches_cols))}201"

    # 4. Reference sheets (read-only)
    ws = wb.create_sheet("Ref_Species")
    write_sheet(ws, species, table_name="tblSpecies")

    # Notes / instructions sheet
    ws = wb.create_sheet("README", 0)
    notes = [
        "WCSAA Catch Entry Template",
        "",
        "How to use:",
        "  1. Add new competitions on the Competitions sheet (comp_id e.g. IC 9).",
        "  2. Add anglers on the Anglers sheet — wp_no must be unique.",
        "  3. Enter every catch on the Catches sheet:",
        "       comp_id (e.g. IC 8) | wp_no (e.g. WP481) | species_raw | length_cm",
        "     Use species_raw exactly as recorded on the angler's catch slip — the",
        "     scoring engine handles aliases, gender suffixes, Site Fish, and < kg variants.",
        "  4. Run: python scripts/score_catches.py    (computes weight, edible, score)",
        "  5. Run: python scripts/generate_reports.py (produces all 7 reports)",
        "",
        "Colour coding:",
        "   Blue text  = manual input (you fill these)",
        "   Black text = computed by scoring engine",
        "",
        "Scoring rules (confirmed 2026-04-28):",
        "  - Weight: W_kg = exp(log_a + b * ln(length_cm))",
        "  - 'Site Fish (...)'  → flat 1.00 kg regardless of length",
        "  - '< X kg' suffix    → score 0 (participation only)",
        "  - Catshark (Brown), Catshark (Puffadder) → score 0",
    ]
    for i, line in enumerate(notes, 1):
        c = ws.cell(row=i, column=1, value=line)
        if i == 1:
            c.font = Font(name=FONT, bold=True, size=14)
        else:
            c.font = Font(name=FONT, size=11)
    ws.column_dimensions["A"].width = 100

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
